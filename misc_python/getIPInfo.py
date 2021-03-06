#!/usr/bin/env python
import socket
import getopt, sys, os
import dns.resolver

try:
    from pygeoip import *
except ImportError:
    print "You must install pygeoip:\n\t https://github.com/appliedsec/pygeoip\n"
    sys.exit(2)

try:
    from openpyxl.workbook import Workbook
    from openpyxl.writer.excel import ExcelWriter
    from openpyxl.cell import get_column_letter
    from openpyxl.styles import Color, Fill, Style, PatternFill, Border, Side, Alignment, Protection, Font
    from openpyxl.cell import Cell
    from openpyxl import load_workbook
    has_openpyxl = True 
except ImportError:
    has_openpyxl = False


'''
Author: Gleeda <jamie.levy@gmail.com>

This program is free software; you can redistribute it and/or
modify it under the terms of the GNU General Public License
as published by the Free Software Foundation; either version
2 of the License, or (at your option) any later version.

getIPInfo.py

 -f <IP file>
 -e <excel output file (default csv to stdout)>
 -g <GeoLiteCity.dat file>
 -c [optional: this will add color to excel file]

What you need:
    1) File with IP addresses, one per line
    2) GeoLiteCity.dat: http://dev.maxmind.com/geoip/legacy/geolite/
    3) Pygeoip: https://github.com/appliedsec/pygeoip
    4) OpenPyxl (optional, for excel output): https://pypi.python.org/pypi/openpyxl
    5) Internet connection for lookups
'''

robtex = "https://www.robtex.com/en/advisory/ip"

codes = {
    "127.0.0.2": "Direct UBE sources, spam operations & spam services",
    "127.0.0.3": "Direct snowshoe spam sources detected via automation",
    "127.0.0.4": "CBL (3rd party exploits such as proxies, trojans, etc.)",
    "127.0.0.5": "CBL (3rd party exploits such as proxies, trojans, etc.)",
    "127.0.0.6": "CBL (3rd party exploits such as proxies, trojans, etc.)",
    "127.0.0.7": "CBL (3rd party exploits such as proxies, trojans, etc.)",
    "127.0.0.10": "End-user Non-MTA IP addresses set by ISP outbound mail policy",
    "127.0.0.11": "End-user Non-MTA IP addresses set by ISP outbound mail policy",
}

servers = [
    "zen.spamhaus.org", 
    "spam.abuse.ch", 
    "virbl.dnsbl.bit.nl", 
    "dnsbl.inps.de", 
    "ix.dnsbl.manitu.net", 
    "dnsbl.sorbs.net", 
    "bl.spamcannibal.org", 
    "bl.spamcop.net", 
    "dnsbl-1.uceprotect.net", 
    "dnsbl-2.uceprotect.net", 
    "dnsbl-3.uceprotect.net", 
    "db.wpbl.info",
    "zen.spamhaus.org",
    "dbl.spamhaus.org",
    "pbl.spamhaus.org",
    "xbl.spamhaus.org",
    "sbl.spamhaus.org",
    "cbl.abuseat.org",
]

if has_openpyxl:
    BoldStyle = Style(font=Font(name='Calibri',
                 size=11,
                 bold=True,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FFFFFFFF'),
            fill=PatternFill(fill_type="solid",
                 start_color='FF000000',
                 end_color='FF000000'))

    RedStyle = Style(font=Font(name='Calibri',
                 size=11,
                 bold=False,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FF000000'),
            border=Border(left=Side(border_style="thick",
                                color='FF000000'),
                      right=Side(border_style="thick",
                                 color='FF000000'),
                      top=Side(border_style="thick",
                               color='FF000000'),
                      bottom=Side(border_style="thick",
                                  color='FF000000'),
                      diagonal=Side(border_style="thick",
                                    color='FF000000'),
                      diagonal_direction=0,
                      outline=Side(border_style="thick",
                                   color='FF000000'),
                      vertical=Side(border_style="thick",
                                    color='FF000000'),
                      horizontal=Side(border_style="thick",
                                     color='FF000000')),
            fill=PatternFill(start_color = 'FFFF0000',
                    end_color = 'FFFF0000',
                    fill_type = 'solid'))

    GreenStyle = Style(font=Font(name='Calibri',
                 size=11,
                 bold=False,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FF000000'),
            border=Border(left=Side(border_style="thin",
                                color='FF000000'),
                      right=Side(border_style="thin",
                                 color='FF000000'),
                      top=Side(border_style="thin",
                               color='FF000000'),
                      bottom=Side(border_style="thin",
                                  color='FF000000'),
                      diagonal=Side(border_style="thin",
                                    color='FF000000'),
                      diagonal_direction=0,
                      outline=Side(border_style="thin",
                                   color='FF000000'),
                      vertical=Side(border_style="thin",
                                    color='FF000000'),
                      horizontal=Side(border_style="thin",
                                     color='FF000000')))
            #fill=PatternFill(start_color = "FF00FF00",
            #        end_color = "FF00FF00",
            #        fill_type = "solid"))

def checkbl(ip):
    for server in servers:
        try:
            my_resolver = dns.resolver.Resolver()
            query = ip + '.' + server
            res = socket.gethostbyname(query)
            answer_txt = "{0}".format(my_resolver.query(query, "TXT")[0] or "")
            return ["Yes", codes[res], server, answer_txt]
        except socket.gaierror:
            pass
    return ["", "", "", ""]

def usage():
    print sys.argv[0], "\n"
    print " -f <IP file>"
    print " -e <excel output file (default csv to stdout)>"
    print " -g <GeoLiteCity.dat file>"
    print " -c [optional: this will add color to excel file]"

def main():
    ipfile = None
    excelfile = None
    GeoLite = None
    wb = None
    ws = None
    color = False
    header = ["IP Address", "Country Code", "Country Name", "Blacklisted", "Code", "Server", "Details", "Robtex Info"]
    try:
        opts, args = getopt.getopt(sys.argv[1:], "hf:g:e:c", ["help", "file=", "geolitecity=", "excelfile=", "coloring"])
    except getopt.GetoptError, err:
        print str(err)
        sys.exit(2)
    
    for o, a in opts:
        if o in ("-h", "--help"):
            usage()
            sys.exit(0)
        elif o in ("-e", "--excelfile"):
            if not has_openpyxl:
                print "You must install OpenPyxl 2.1.2+ for xlsx format:\n\thttps://pypi.python.org/pypi/openpyxl"
                sys.exit(-1)
            excelfile = a
            wb = Workbook(optimized_write = True)
            ws = wb.create_sheet()
            ws.title = "IP Address Info"
            ws.append(header)
        elif o in ("-g", "--geolitecity"):
            GeoLite = a
        elif o in ("-f", "--file"):
            if os.path.isfile(a):
                ipfile = open(a, "r")
            else:
                print a + " is not a file"
                usage()
                sys.exit(-1)
        elif o in ("-c", "--coloring"):
            color = True
        else:
            assert False, "unhandled option\n\n"
            sys.exit(-2)

    if ipfile == None or GeoLite == None:
        usage()
        sys.exit(2)

    gi = GeoIP(GeoLite)
    if excelfile == None:
        print ",".join(x for x in header)

    total = 1
    for i in ipfile.readlines():
        item =  i.strip()
        rev = '.'.join(item.split('.')[::-1])
        gistuff = gi.record_by_addr(item)
        ip = "/".join(item.split("."))
        robtex_url = "{0}/{1}".format(robtex, ip)
        bl = checkbl(rev)
        line = [item, gistuff["country_code"], gistuff["country_name"], bl[0], bl[1], bl[2], bl[3], robtex_url]
        if excelfile == None:
            print ",".join(x for x in line)
        else:
            ws.append(line)
            total += 1

    if excelfile:
        wb.save(filename = excelfile)
        if not color:
            return
        wb = load_workbook(filename = excelfile)
        ws = wb.get_sheet_by_name(name = "IP Address Info")
        for col in xrange(1, len(header) + 1):
            ws.cell("{0}{1}".format(get_column_letter(col), 1)).style = BoldStyle

        for row in xrange(2, total + 1):
            for col in xrange(1, len(header) + 1):
                if ws.cell("{0}{1}".format(get_column_letter(4), row)).value == "Yes":
                    ws.cell("{0}{1}".format(get_column_letter(col), row)).style = RedStyle
                else:
                    ws.cell("{0}{1}".format(get_column_letter(col), row)).style = GreenStyle
        wb.save(filename = excelfile)

if __name__ == "__main__":
    main()
